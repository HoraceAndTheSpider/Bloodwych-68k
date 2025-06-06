#!/usr/bin/env python3
import os
import sys
import pandas as pd
import re
import shutil
from openpyxl import load_workbook

from .tool_common import parse_int

# 3. Inspect segments (read-only)
def inspect_source(master, sheet, name_filter=None, label_filter=None, debug=False):
    """
    Scan each segment from the spreadsheet, verify that its
    dc.* block in ASM matches the extracted .bin file head/tail bytes,
    then prepare to replace each matching block with an INCBIN directive.

    Args:
      master        – the base name of the master binary (e.g. "BLOODWYCH439")
      sheet         – path to the Excel/CSV sheet
      name_filter   – if provided, only segments whose 'name' exactly matches (case-insensitive) are checked
      label_filter  – if provided, only segments whose ASM label (or fallback original label) exactly matches (case-insensitive) are checked
      debug         – if True, print head/tail differences on failure
    """
    base = os.path.splitext(master)[0]
    clean_dir = os.path.join('data', f"{base}-clean")

    # Load the workbook and pick the correct sheet
    wb = load_workbook(sheet, data_only=True)
    sheet_name = next((s for s in wb.sheetnames if s.lower() == master.lower()), None)
    if not sheet_name:
        print(f"No sheet {master}")
        sys.exit(1)

    # Decide whether we're using <master>_relabel.asm or <master>.asm
    asm_folder = 'asm'
    relabel_f = os.path.join(asm_folder, f"{master}_relabel.asm")
    if os.path.isfile(relabel_f):
        asm_path, label_col = relabel_f, 'relabel'
        print(f"Using relabel ASM {asm_path}")
    else:
        asm_path, label_col = os.path.join(asm_folder, f"{master}.asm"), 'label'
    if not os.path.isfile(asm_path):
        print(f"No ASM {asm_path}")
        sys.exit(1)

    # Read the original ASM lines into memory, and make a working copy
    lines = open(asm_path, 'r', errors='ignore').read().splitlines()
    working_lines = list(lines)

    # Load the spreadsheet data into a DataFrame
    df = pd.read_excel(sheet, sheet_name=sheet_name)
    df.columns = [c.strip().lower() for c in df.columns]
    for req in ('name', 'size', label_col):
        if req not in df.columns:
            print(f"Missing {req}")
            sys.exit(1)

    # We'll collect all (start, end, seg_path) tuples for later replacement
    mods = []

    for idx, r in df.iterrows():
        row = idx + 2  # for user-friendly reporting (“row 2” corresponds to Excel’s row 2)
        nm = str(r['name']).strip()
        lb = str(r[label_col]).strip()
        if not nm or nm.lower() == 'nan' or not lb or lb.lower() == 'nan':
            continue

        # 1) Apply the 'name_filter' if given
        if name_filter and nm.lower() != name_filter.lower():
            continue

        # 2) Apply the 'label_filter' if given:
        #    If we're in relabel mode, we also allow 'fallback_label' below to match
        if label_filter:
            # Compare against the relabel column first (lb), then if label_col=='relabel',
            # compare against the original 'label' column as fallback.
            if lb.lower() != label_filter.lower():
                if label_col == 'relabel':
                    orig_lbl = str(r.get('label', '')).strip()
                    if not orig_lbl or orig_lbl.lower() != label_filter.lower():
                        continue
                else:
                    continue

        sz = parse_int(r['size'])
        print(f"Check {nm}@{lb} row{row}")
        seg_path = os.path.join(clean_dir, nm)
        if sz is None or not os.path.isfile(seg_path) or os.path.getsize(seg_path) != sz:
            print("  missing or size mismatch")
            continue

        # 3) Locate the label block in the ASM: try relabel first, then fallback to original 'label'
        relb = lb  # the value from r[label_col]
        start = None

        # (1) Try matching relabel (i.e. lb) exactly
        if relb:
            start = next((i for i, line in enumerate(lines)
                          if line.lower().startswith(relb.lower() + ':')), None)

        # (2) If that failed AND we're in relabel mode, try the original 'label' as fallback
        if start is None and label_col == 'relabel':
            orig_lbl = str(r.get('label', '')).strip()
            if orig_lbl:
                start = next((i for i, line in enumerate(lines)
                              if line.lower().startswith(orig_lbl.lower() + ':')), None)
                if start is not None:
                    print(f"  falling back to original label '{orig_lbl}'")

        # (3) If still not found, skip
        if start is None:
            print(f"  no label match for '{nm}'; skipping")
            continue

        # 4) Find the end of this declared data block: next bare “something:” or end-of-file
        end = next((i for i in range(start + 1, len(lines))
                    if lines[i].strip().endswith(':')), len(lines))

        # 5) Read the extracted binary and compare head/tail against in-source dc.* directives
        data = open(seg_path, 'rb').read()
        chk = min(16, sz)
        head = data[:chk]
        tail = data[-chk:]

        dh = bytearray()
        dt = bytearray()

        # Build 'dh' forward from the dc.* lines
        for ln in lines[start + 1:end]:
            txt = ln.strip()
            if not txt or txt.startswith(';') or txt.endswith(':'):
                if txt.endswith(':'):
                    break
                continue

            m = re.match(r"\s*dc\.(b|w|l)\s*(.+)", ln, re.I)
            if not m:
                break

            k = m.group(1).lower()  # 'b', 'w', or 'l'
            its = m.group(2)
            for part in its.split(','):
                raw = part.split(';')[0].strip()
                # Handle string literals: 'HELLO' or "WORLD"
                if (raw.startswith("'") and raw.endswith("'")) or (raw.startswith('"') and raw.endswith('"')):
                    for c in raw[1:-1]:
                        dh.append(ord(c))
                    continue

                if raw.startswith('$'):
                    raw = '0x' + raw[1:]
                try:
                    v = int(raw, 16)
                except ValueError:
                    continue

                if k == 'b':
                    dh.append(v & 0xFF)
                elif k == 'w':
                    dh.extend([ (v >> 8) & 0xFF, v & 0xFF ])
                else:  # 'l'
                    dh.extend([
                        (v >> 24) & 0xFF,
                        (v >> 16) & 0xFF,
                        (v >> 8) & 0xFF,
                        v & 0xFF
                    ])
            if len(dh) >= chk:
                break

        # Build 'dt' backward from the dc.* lines
        for ln in reversed(lines[start + 1:end]):
            txt = ln.strip()
            if not txt or txt.startswith(';') or txt.endswith(':'):
                if txt.endswith(':'):
                    break
                continue

            m = re.match(r"\s*dc\.(b|w|l)\s*(.+)", ln, re.I)
            if not m:
                break

            k = m.group(1).lower()
            parts = [p.split(';')[0].strip() for p in m.group(2).split(',')]
            for raw in reversed(parts):
                # Handle string literal trailing in this block
                if (raw.startswith("'") and raw.endswith("'")) or (raw.startswith('"') and raw.endswith('"')):
                    blk = [ord(c) for c in raw[1:-1]]
                    dt[0:0] = blk
                    continue

                if raw.startswith('$'):
                    raw = '0x' + raw[1:]
                try:
                    v = int(raw, 16)
                except ValueError:
                    continue

                if k == 'b':
                    blk = [v & 0xFF]
                elif k == 'w':
                    blk = [ (v >> 8) & 0xFF, v & 0xFF ]
                else:
                    blk = [
                        (v >> 24) & 0xFF,
                        (v >> 16) & 0xFF,
                        (v >> 8) & 0xFF,
                        v & 0xFF
                    ]
                dt[0:0] = blk  # prepend this block
            if len(dt) >= chk:
                break

        # Compare expected vs. in-source bytes
        if head == bytes(dh[:chk]) and tail == bytes(dt[-chk:]):
            print("  VALID")
        else:
            print("  FAIL")
            if debug:
                print(f"    expected head: {head.hex()}")
                print(f"    found    head: {bytes(dh[:chk]).hex()}")
                print(f"    expected tail: {tail.hex()}")
                print(f"    found    tail: {bytes(dt[-chk:]).hex()}")

        # Schedule this block for replacement with INCBIN later
        mods.append((start, end, os.path.join(clean_dir, nm)))

    # Apply all replacements in _reverse_ order so earlier edits don't shift later indices
    for start, end, seg_path in sorted(mods, key=lambda x: x[0], reverse=True):
        # Delete the old dc.* lines
        del working_lines[start + 1 : end]
        # Insert the new INCBIN directive (with a leading forward slash)
        working_lines.insert(start + 1, f'\tINCBIN "/{seg_path}"\n')

    # Write out the modified ASM file under *_data.asm
    base_name, ext = os.path.splitext(os.path.basename(asm_path))
    new_name = os.path.join(os.path.dirname(asm_path), f"{base_name}_data{ext}")
    with open(new_name, 'w') as fout:
        fout.write("\n".join(working_lines))

    print(f"Modified ASM written to {new_name}")
