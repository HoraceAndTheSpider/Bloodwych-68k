#!/usr/bin/env python3
import os
import sys
import argparse
import pandas as pd
import re
import shutil
from openpyxl import load_workbook

# Common utility to parse integers
def parse_int(value):
    try:
        if pd.isna(value): return None
        if isinstance(value, str):
            v = value.strip()
            if not v: return None
            if v.lower().startswith('0x') or v.startswith('$'):
                return int(v.replace('$','0x'), 16)
            return int(v)
        return int(value)
    except:
        return None

# 1. Extract segments
def extract_segments(master, sheet, name_filter=None, debug=False):
    binaries = 'binaries'
    master_path = os.path.join(binaries, master)
    if not os.path.isdir(binaries) or not os.path.isfile(master_path):
        print(f"Error: master binary '{master}' not found in '{binaries}/'")
        sys.exit(1)
    base, _ = os.path.splitext(master)
    out_dir = f"extracted-{base}"
    os.makedirs(out_dir, exist_ok=True)

    df = pd.read_excel(sheet) if sheet.lower().endswith(('.xls','.xlsx')) else pd.read_csv(sheet)
    df.columns = [c.strip().lower() for c in df.columns]
    for col in ('offset','size','name'):
        if col not in df.columns:
            print(f"Missing column '{col}' in sheet")
            sys.exit(1)

    with open(master_path, 'rb') as mf:
        for _, row in df.iterrows():
            nm = str(row['name']).strip()
            if not nm or pd.isna(row['name']):
                continue
            if name_filter and nm.lower() != name_filter.lower():
                continue
            off = parse_int(row['offset'])
            sz  = parse_int(row['size'])
            if off is None or sz is None:
                if debug: print(f"Skipping '{nm}': invalid offset/size")
                continue
            mf.seek(off)
            data = mf.read(sz)
            out_path = os.path.join(out_dir, nm)
            os.makedirs(os.path.dirname(out_path), exist_ok=True)
            with open(out_path, 'wb') as f:
                f.write(data)
            print(f"Extracted '{nm}'")

# 2. Patch segments
def patch_segments(master, sheet, name_filter=None, debug=False):
    binaries = 'binaries'
    orig = os.path.join(binaries, master)
    if not os.path.isdir(binaries) or not os.path.isfile(orig):
        print(f"Error: master binary '{master}' not found in '{binaries}/'")
        sys.exit(1)
    base, ext = os.path.splitext(master)
    mod_dir = f"modified-{base}"
    patched = os.path.join(binaries, f"{base}-modified{ext}")
    shutil.copy2(orig, patched)
    print(f"Patching into '{patched}'")

    df = pd.read_excel(sheet) if sheet.lower().endswith(('.xls','.xlsx')) else pd.read_csv(sheet)
    df.columns = [c.strip().lower() for c in df.columns]
    for col in ('offset','size','name'):
        if col not in df.columns:
            print(f"Missing column '{col}' in sheet")
            sys.exit(1)

    with open(patched, 'r+b') as mf:
        for _, row in df.iterrows():
            nm = str(row['name']).strip()
            if not nm or pd.isna(row['name']):
                continue
            if name_filter and nm.lower() != name_filter.lower():
                continue
            off = parse_int(row['offset'])
            sz  = parse_int(row['size'])
            if off is None or sz is None:
                if debug: print(f"Skipping '{nm}': invalid offset/size")
                continue
            seg = os.path.join(mod_dir, nm)
            if not os.path.isfile(seg) or os.path.getsize(seg) != sz:
                if debug: print(f"Skipping '{nm}': file missing or size mismatch")
                continue
            data = open(seg,'rb').read()
            mf.seek(off)
            mf.write(data)
            print(f"Patched '{nm}'")

# 3. Inspect segments (read-only)
def inspect_source(master, sheet, name_filter=None, debug=False):
    import pandas as pd
    from openpyxl import load_workbook

    # 1) Pick worksheet
    wb_val = load_workbook(sheet, data_only=True)
    sheet_name = next((s for s in wb_val.sheetnames if s.lower() == master.lower()), None)
    if not sheet_name:
        print(f"Error: worksheet '{master}' not found in workbook")
        sys.exit(1)

    # 2) Pick ASM, relabel or original, and label column
    asm_folder = 'asm'
    relabel_asm = os.path.join(asm_folder, f"{master}_relabel.asm")
    if os.path.isfile(relabel_asm):
        asm_path, label_col = relabel_asm, 'relabel'
        print(f"Using relabeled ASM: {asm_path}")
    else:
        asm_path, label_col = os.path.join(asm_folder, f"{master}.asm"), 'label'
    if not os.path.isfile(asm_path):
        print(f"Error: ASM '{asm_path}' not found")
        sys.exit(1)

    # 3) Load sheet via pandas
    df = pd.read_excel(sheet, sheet_name=sheet_name)
    df.columns = [str(c).strip().lower() for c in df.columns]
    for req in ('name', 'size', label_col):
        if req not in df.columns:
            print(f"Error: '{req}' column required")
            sys.exit(1)

    # 4) Read ASM lines
    asm_lines = open(asm_path, 'r', encoding='utf-8', errors='ignore').read().splitlines()

    # 5) Iterate
    for idx, row in df.iterrows():
        excel_row = idx + 2
        name  = str(row['name']).strip()
        if not name or name.lower()=='nan': continue
        label = str(row[label_col]).strip()
        if not label or label.lower()=='nan': continue
        if name_filter and name.lower()!=name_filter.lower(): continue

        size = parse_int(row['size'])
        print(f"Checking '{name}' at label '{label}' (row {excel_row})")

        seg_path = os.path.join(f"extracted-{master}", name)
        if size is None or not os.path.isfile(seg_path) or os.path.getsize(seg_path)!=size:
            print("  file missing or size mismatch")
            continue
        data = open(seg_path,'rb').read()

        start_idx = next((i for i,l in enumerate(asm_lines) if l.lower().startswith(label.lower()+':')), None)
        if start_idx is None:
            print("  label not found in ASM")
            continue
        end_idx = next((i for i in range(start_idx+1,len(asm_lines)) if asm_lines[i].strip().endswith(':')), len(asm_lines))

        chk = min(16, size)
        head, tail = data[:chk], data[-chk:]

        # Build head (skip blanks/comments, stop on labels/non-dc)
        dh = bytearray()
        for ln in asm_lines[start_idx+1:end_idx]:
            if not ln.strip() or ln.strip().startswith(';') or ln.strip().endswith(':'):
                if ln.strip().endswith(':'): break
                continue
            m = re.match(r'\s*dc\.(b|w|l)\s*(.+)', ln, re.I)
            if not m: break
            kind, items = m.group(1).lower(), m.group(2)
            for part in items.split(','):
                raw = part.split(';')[0].strip()
                if raw.startswith('$'): raw = '0x'+raw[1:]
                try: v=int(raw,16)
                except: continue
                if kind=='b': dh.append(v&0xFF)
                elif kind=='w': dh.extend([(v>>8)&0xFF,v&0xFF])
                else: dh.extend([(v>>24)&0xFF,(v>>16)&0xFF,(v>>8)&0xFF,v&0xFF])
            if len(dh)>=chk: break

        # Build tail similarly
        dt = bytearray()
        for ln in reversed(asm_lines[start_idx+1:end_idx]):
            if not ln.strip() or ln.strip().startswith(';') or ln.strip().endswith(':'):
                if ln.strip().endswith(':'): break
                continue
            m = re.match(r'\s*dc\.(b|w|l)\s*(.+)', ln, re.I)
            if not m: break
            kind, items = m.group(1).lower(), m.group(2)
            parts = [p.split(';')[0].strip() for p in items.split(',')]
            for raw in reversed(parts):
                if raw.startswith('$'): raw='0x'+raw[1:]
                try: v=int(raw,16)
                except: continue
                blk = ([v&0xFF] if kind=='b' else([(v>>8)&0xFF,v&0xFF] if kind=='w' else[(v>>24)&0xFF,(v>>16)&0xFF,(v>>8)&0xFF,v&0xFF]))
                dt[0:0]=blk
            if len(dt)>=chk: break

        # Validate
        if head==bytes(dh[:chk]) and tail==bytes(dt[-chk:]):
            print("  VALIDATED in source")
        else:
            print("  VALIDATION FAILED")
            if debug:
                print(f"    head      : {head.hex()}")
                print(f"    decl_head : {bytes(dh[:chk]).hex()}")
                print(f"    tail      : {tail.hex()}")
                print(f"    decl_tail : {bytes(dt[-chk:]).hex()}")

    # 6) Save data-check copy
    base, ext = os.path.splitext(os.path.basename(asm_path))
    data_name   = f"{base}_data{ext}"
    data_path   = os.path.join(os.path.dirname(asm_path), data_name)
    shutil.copy2(asm_path, data_path)
    print(f"Saved data-check ASM copy to '{data_path}'")
    print("Done checking—all results printed above.")

# 4. Relabel segments in ASM source
#    Adds a '_relabel' copy with label renames based on 'label' and 'relabel' columns

def relabel_segments(master, sheet):
    # Load relabel instructions
    df = pd.read_excel(sheet) if sheet.lower().endswith(('.xls','.xlsx')) else pd.read_csv(sheet)
    df.columns = [c.strip().lower() for c in df.columns]
    # Ensure necessary columns
    for req in ('label', 'relabel'):
        if req not in df.columns:
            print(f"Error: '{req}' column required for relabel")
            sys.exit(1)

    # Prepare ASM files
    asm_folder = 'asm'
    orig_path = os.path.join(asm_folder, f"{master}.asm")
    if not os.path.isfile(orig_path):
        print(f"Error: ASM '{orig_path}' not found")
        sys.exit(1)
    base, ext = os.path.splitext(master)
    new_name = f"{base}_relabel{ext}.asm"
    new_path = os.path.join(asm_folder, new_name)
    shutil.copy2(orig_path, new_path)
    print(f"Created relabel copy '{new_path}'")

    # Read content lines
    with open(new_path, 'r', encoding='utf-8', errors='ignore') as f:
        lines = f.read().splitlines()

    # Process each relabel instruction
    for _, row in df.iterrows():
        label = str(row['label']).strip()
        new_label = str(row['relabel']).strip()
        if not label or label.lower() == 'nan':
            continue
        if not new_label or new_label.lower() == 'nan' or new_label == label:
            continue

        # 1) _delete directives
        if new_label.startswith('_delete'):
            # Find lines defining the label
            matches = [i for i, ln in enumerate(lines)
                       if re.match(rf'^\s*{re.escape(label)}\s*:', ln)]
            if len(matches) == 1:
                idx = matches[0]
                line = lines[idx]
                # Only delete if there's no comment
                if ';' not in line:
                    print(f"Deleting label line '{label}:' at line {idx+1}")
                    lines.pop(idx)
                else:
                    print(f"Cannot delete '{label}': comment present on line {idx+1}")
            else:
                print(f"Cannot delete '{label}': found {len(matches)} occurrences")
            continue

        # 2) _offset directives (placeholder)
        if new_label.lower().startswith('_offset_'):
            parts = new_label.split('_')
            if len(parts) < 4 or not parts[-1].lower().startswith('0x'):
                print(f"Invalid _offset format '{new_label}', skipping")
                continue
            # Join all name parts between the first and the last element
            name_parts = parts[2:-1]
            name_part = '_'.join(name_parts).rstrip('_')
            hex_part    = parts[-1]
            offset_lbl  = f"{name_part}+${hex_part[2:]}"

            # remove the label definition
            matches = [i for i, ln in enumerate(lines)
                       if re.match(rf'^\s*{re.escape(label)}\s*:', ln)]
            if len(matches) == 1:
                idx = matches[0]
                if ';' not in lines[idx]:
                    print(f"Deleting label line '{label}:' at line {idx+1}")
                    lines.pop(idx)
                else:
                    print(f"Cannot delete '{label}': comment present on line {idx+1}")
            else:
                print(f"Cannot delete '{label}': found {len(matches)} occurrences")

            # patch all instruction references
            for i, ln in enumerate(lines):
                if re.match(rf'^\s*{re.escape(label)}\s*:', ln):
                    continue
                if re.search(rf'\b{re.escape(label)}\b', ln):
                    new_ln = re.sub(rf'\b{re.escape(label)}\b', offset_lbl, ln)
                    lines[i] = new_ln
                    print(f"Replaced '{label}' -> '{offset_lbl}' in line {i+1}")
            continue

        # 3) General relabel
        # Verify the definition exists
        if not any(re.match(rf'^\s*{re.escape(label)}\s*:', ln) for ln in lines):
            print(f"Label '{label}' not found in source, skipping")
            continue
        # Replace all whole-word occurrences
        pattern = rf'\b{re.escape(label)}\b'
        lines = [re.sub(pattern, new_label, ln) for ln in lines]
        print(f"Relabeled '{label}' to '{new_label}'")

    # Save updated ASM
    with open(new_path, 'w', encoding='utf-8', errors='ignore') as f:
        f.write("\n".join(lines) + "\n")
    print(f"Saved relabeled ASM to '{new_path}'")

# Entry point
def main():
    parser = argparse.ArgumentParser(description="Segment tool: extract, patch, inspect, relabel")
    sub = parser.add_subparsers(dest='cmd')
    extract_p = sub.add_parser('extract', help='Extract segments')
    extract_p.add_argument('-n','--name', help='Filter segment name')
    extract_p.add_argument('--debug', action='store_true', help='Enable debug output')
    patch_p   = sub.add_parser('patch', help='Patch segments')
    patch_p.add_argument('-n','--name', help='Filter segment name')
    patch_p.add_argument('--debug', action='store_true', help='Enable debug output')
    inspect_p = sub.add_parser('inspect', help='Inspect segments')
    inspect_p.add_argument('-n','--name', help='Filter segment name')
    inspect_p.add_argument('--debug', action='store_true', help='Show head/tail bytes on failure')
    relabel_p = sub.add_parser('relabel', help='Relabel labels in ASM source')
    # No -n filter for relabel; it processes all
    parser.add_argument('-m', '--master', default='BLOODWYCH439', help='Master binary name')
    parser.add_argument('-s', '--sheet',  default='segments.xlsx', help='Spreadsheet file')
    args = parser.parse_args()

    if args.cmd == 'extract':
        extract_segments(args.master, args.sheet, name_filter=getattr(args,'name',None), debug=getattr(args,'debug',False))
    elif args.cmd == 'patch':
        patch_segments(args.master, args.sheet, name_filter=getattr(args,'name',None), debug=getattr(args,'debug',False))
    elif args.cmd == 'inspect':
        inspect_source(args.master, args.sheet, name_filter=getattr(args,'name',None), debug=getattr(args,'debug',False))
    elif args.cmd == 'relabel':
        relabel_segments(args.master, args.sheet)
    else:
        parser.print_help()

if __name__=='__main__':
    main()
