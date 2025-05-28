import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import sys
import os.path


# handle the different binarys - bloodywch, extended levels and ST demo.
ThisGame = "bw"
# ThisGame="bext"
# ThisGame="demo"

if ThisGame == "bw":
    ThisBinary = "bw439"
    
elif ThisGame == "bext":
    ThisBinary = "bext43"

elif ThisGame == "demo":
    ThisBinary = "AtariST_DEMO_CODE"

# do some checks that the binary we want is there.
binpath = "../../bin/"
binfullpath = os.path.join(binpath , ThisBinary)

if os.path.isfile(binfullpath) == False:
    print ("No file found: " & binfullpath)
    sys.exit()
    
    
# open the game binary
game = open(binfullpath,"rb")



# set a workbook object
wb = Workbook()

# load the XLSM and pick the desired tab
wb = load_workbook("../../Data Mapper.xlsm",data_only=True)
ws = wb["BWDataExtract"]


# iterate through the sheet
for rowcount in range(2,200):

    # From XLSM get the name and type of block
    category = (ws.cell(row=rowcount, column=1).value)
    name = (ws.cell(row=rowcount, column=2).value)

    #print (name)
    # From XLSM get the position of the data, exclude blanks
    if (ws.cell(row=rowcount, column=5).value != None):
        position= int(ws.cell(row=rowcount, column=5).value)
    else:
        position = 0        

    # From XLSM get the length of the data, exclude blanks
    if (ws.cell(row=rowcount, column=6).value != None ):
       # print (ws.cell(row=rowcount, column=6).value)    
        length = int(ws.cell(row=rowcount, column=6).value)
    else:
        length = 0

    # check if the category folder exists
    filepath = os.path.join("../../"+ThisGame+"-data-clean" , category)
    if os.path.isdir(filepath) == False:
        os.mkdir (filepath)
        print("Creating path... " + filepath)

    
    # we will ignore any blanks or zero length
    if (category != "" and category != "."\
        and name !="" and name != "."\
        and length >0 and position!=0 ):

        # jump to the right point in the game binary
        game.seek(position,0)

        # creat our output filename and binary block
        filename = os.path.join(filepath , name)
        savefile = open(filename,"wb")

        # read the data
        block = game.read(length)

        # save it out        
        savefile.write(block)
        print ("Saving file... " + category +"\\"+ name)
        savefile.close()

       # sys.exit()

game.close()       

